from __future__ import annotations

import contextlib
import re
import tempfile
import time
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from itertools import count, zip_longest
from typing import Any, BinaryIO
from uuid import uuid4
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from snax_import.domain.raw_workbook import (
    Cell,
    CellCoordinate,
    FilenameMetadata,
    Formula,
    MergedRange,
    RawValue,
    Row,
    Sheet,
    SheetVisibility,
    ValueType,
    Workbook,
    WorkbookFormat,
)
from snax_import.ports.workbook_reader import (
    IssueSeverity,
    ReaderIssue,
    ReaderIssueCode,
    ReaderOptions,
    ReaderResult,
    ReaderStatistics,
    WorkbookReader,
)

_CHUNK_SIZE = 1024 * 1024
_SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx"}
_SUPPORTED_MEDIA_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.template",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
}
_FORMULA_ERROR_CODE_BY_TOKEN = {
    "#NULL!": "FORMULA_ERROR_NULL",
    "#DIV/0!": "FORMULA_ERROR_DIV_ZERO",
    "#VALUE!": "FORMULA_ERROR_VALUE",
    "#REF!": "FORMULA_ERROR_REF",
    "#NAME?": "FORMULA_ERROR_NAME",
    "#NUM!": "FORMULA_ERROR_NUM",
    "#N/A": "FORMULA_ERROR_NA",
    "#GETTING_DATA": "FORMULA_ERROR_GETTING_DATA",
    "#SPILL!": "FORMULA_ERROR_SPILL",
    "#FIELD!": "FORMULA_ERROR_FIELD",
    "#CALC!": "FORMULA_ERROR_CALC",
}


_CELL_REFERENCE_PATTERN = re.compile(r"^(?P<column>[A-Z]+)(?P<row>\d+)$")


@dataclass(frozen=True, slots=True)
class _ParsedRowResult:
    row: Row | None
    parsed_cells: int
    formula_cells: int
    error_cells: int
    stopped: bool = False


@dataclass(frozen=True, slots=True)
class _ParsedSheetResult:
    sheet: Sheet
    rows_read: int
    rows_skipped: int
    cells_read: int
    formula_cells: int
    error_cells: int
    stopped: bool = False


@dataclass(frozen=True, slots=True)
class _ParsedWorkbookResult:
    workbook: Workbook | None
    sheets_read: int
    rows_read: int
    cells_read: int
    formula_cells: int
    error_cells: int
    skipped_sheets: int
    skipped_rows: int


class XlsxWorkbookReader(WorkbookReader):
    """Read-only XLSX parser that stores formulas as raw text with cached values."""

    def supports(self, media_type: str | None = None, extension: str | None = None) -> bool:
        media_supported = (
            media_type is not None and media_type.lower().split(";", 1)[0] in _SUPPORTED_MEDIA_TYPES
        )
        extension_supported = extension is not None and extension.lower() in _SUPPORTED_EXTENSIONS
        return (media_type is not None or extension is not None) and (
            media_supported or extension_supported
        )

    def read(self, source: BinaryIO, options: ReaderOptions) -> ReaderResult:
        started = time.monotonic()
        issue_counter = count(1)
        issues: list[ReaderIssue] = []

        def add_issue(
            code: ReaderIssueCode,
            severity: IssueSeverity,
            message: str,
            *,
            sheet_name: str | None = None,
            row_index: int | None = None,
            cell_coordinate: CellCoordinate | None = None,
            retryable: bool = False,
            details: dict[str, str] | None = None,
        ) -> None:
            issues.append(
                ReaderIssue(
                    issue_id=f"xlsx-reader-{next(issue_counter):06d}",
                    code=code,
                    severity=severity,
                    message=message,
                    sheet_name=sheet_name,
                    row_index=row_index,
                    cell_coordinate=cell_coordinate,
                    retryable=retryable,
                    details=details or {},
                )
            )

        bytes_read = 0
        parsed: _ParsedWorkbookResult
        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except Exception as exc:
            add_issue(
                ReaderIssueCode.UNSUPPORTED_FORMAT,
                IssueSeverity.ERROR,
                "openpyxl runtime dependency is not available",
                details={"error": type(exc).__name__},
            )
            return ReaderResult(
                workbook=None,
                issues=tuple(issues),
                statistics=self._statistics(bytes_read=0, started=started),
            )

        try:
            with tempfile.TemporaryFile() as temp:
                copied = self._copy_source_to_temp(
                    source=source,
                    destination=temp,
                    options=options,
                    started=started,
                    add_issue=add_issue,
                )
                if copied is None:
                    return ReaderResult(
                        workbook=None,
                        issues=tuple(issues),
                        statistics=self._statistics(bytes_read=0, started=started),
                    )
                bytes_read = copied
                temp.seek(0)
                parsed = self._parse_workbook(
                    source=temp,
                    source_size=bytes_read,
                    load_workbook=load_workbook,
                    options=options,
                    started=started,
                    add_issue=add_issue,
                )
        except Exception as exc:
            add_issue(
                ReaderIssueCode.UNSUPPORTED_FORMAT,
                IssueSeverity.ERROR,
                "Failed to parse XLSX workbook",
                details={"error": type(exc).__name__},
            )
            parsed = _ParsedWorkbookResult(
                workbook=None,
                sheets_read=0,
                rows_read=0,
                cells_read=0,
                formula_cells=0,
                error_cells=0,
                skipped_sheets=0,
                skipped_rows=0,
            )

        statistics = ReaderStatistics(
            sheets_read=parsed.sheets_read,
            rows_read=parsed.rows_read,
            cells_read=parsed.cells_read,
            formula_cells=parsed.formula_cells,
            error_cells=parsed.error_cells,
            skipped_sheets=parsed.skipped_sheets,
            skipped_rows=parsed.skipped_rows,
            bytes_read=bytes_read,
            duration_seconds=self._elapsed(started),
        )
        return ReaderResult(workbook=parsed.workbook, issues=tuple(issues), statistics=statistics)

    def _parse_workbook(
        self,
        *,
        source: Any,
        source_size: int,
        load_workbook: Any,
        options: ReaderOptions,
        started: float,
        add_issue: Any,
    ) -> _ParsedWorkbookResult:
        source.seek(0)
        merged_ranges_by_sheet = self._collect_merged_ranges(source=source)
        source.seek(0)
        formula_book = load_workbook(source, read_only=True, data_only=False, keep_links=False)
        source.seek(0)
        cached_book = load_workbook(source, read_only=True, data_only=True, keep_links=False)
        source.seek(0)

        with contextlib.ExitStack() as stack:
            stack.callback(formula_book.close)
            stack.callback(cached_book.close)

            total_sheets = max(len(formula_book.worksheets), len(cached_book.worksheets))
            sheets: list[Sheet] = []
            rows_read = 0
            cells_read = 0
            formula_cells = 0
            error_cells = 0
            skipped_sheets = 0
            skipped_rows = 0

            for sheet_index in range(total_sheets):
                formula_sheet = (
                    formula_book.worksheets[sheet_index]
                    if sheet_index < len(formula_book.worksheets)
                    else None
                )
                cached_sheet = (
                    cached_book.worksheets[sheet_index]
                    if sheet_index < len(cached_book.worksheets)
                    else None
                )
                if formula_sheet is None or cached_sheet is None:
                    add_issue(
                        ReaderIssueCode.SHEET_READ_FAILED,
                        IssueSeverity.ERROR,
                        "Workbook worksheets are structurally inconsistent",
                        details={"sheetIndex": str(sheet_index)},
                    )
                    skipped_sheets += 1
                    continue

                if self._is_timeout(started, options):
                    add_issue(
                        ReaderIssueCode.TIMEOUT_EXCEEDED,
                        IssueSeverity.CRITICAL,
                        "XLSX reader timeout exceeded",
                        retryable=True,
                    )
                    skipped_sheets += total_sheets - sheet_index
                    break

                if sheet_index >= options.max_sheets:
                    add_issue(
                        ReaderIssueCode.WORKBOOK_TOO_MANY_SHEETS,
                        IssueSeverity.CRITICAL,
                        "Workbook exceeds max_sheets",
                        details={"limit": str(options.max_sheets)},
                    )
                    skipped_sheets += total_sheets - sheet_index
                    break

                if rows_read >= options.max_rows:
                    add_issue(
                        ReaderIssueCode.WORKBOOK_TOO_MANY_ROWS,
                        IssueSeverity.CRITICAL,
                        "Workbook exceeds max_rows",
                        details={"limit": str(options.max_rows)},
                    )
                    skipped_sheets += total_sheets - sheet_index
                    break

                parsed_sheet = self._parse_sheet(
                    sheet=formula_sheet,
                    cached_sheet=cached_sheet,
                    sheet_index=sheet_index,
                    rows_read=rows_read,
                    cells_read=cells_read,
                    merged_ranges_by_sheet=merged_ranges_by_sheet,
                    options=options,
                    started=started,
                    add_issue=add_issue,
                )

                sheets.append(parsed_sheet.sheet)
                rows_read += parsed_sheet.rows_read
                skipped_rows += parsed_sheet.rows_skipped
                cells_read += parsed_sheet.cells_read
                formula_cells += parsed_sheet.formula_cells
                error_cells += parsed_sheet.error_cells

                if parsed_sheet.stopped:
                    break

        return _ParsedWorkbookResult(
            workbook=Workbook(
                id=uuid4(),
                source_file_id=uuid4(),
                filename=FilenameMetadata(
                    name="workbook.xlsx",
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    size_bytes=source_size,
                ),
                format=WorkbookFormat.XLSX,
                created_at=datetime.now(UTC),
                sheets=tuple(sheets),
                workbook_metadata={"parser": "openpyxl"},
            ),
            sheets_read=len(sheets),
            rows_read=rows_read,
            cells_read=cells_read,
            formula_cells=formula_cells,
            error_cells=error_cells,
            skipped_sheets=skipped_sheets,
            skipped_rows=skipped_rows,
        )

    def _parse_sheet(
        self,
        *,
        sheet: Any,
        cached_sheet: Any,
        sheet_index: int,
        rows_read: int,
        cells_read: int,
        merged_ranges_by_sheet: dict[str, tuple[str, ...]],
        options: ReaderOptions,
        started: float,
        add_issue: Any,
    ) -> _ParsedSheetResult:
        sheet_name = str(sheet.title)
        visibility = self._sheet_visibility(str(getattr(sheet, "sheet_state", "visible")))
        max_row = int(sheet.max_row or 0)
        max_column = int(sheet.max_column or 0)
        remaining_rows = max(0, options.max_rows - rows_read)
        rows_to_read = min(max_row, remaining_rows)
        merged_ranges = self._parse_merged_ranges(
            sheet=sheet,
            sheet_name=sheet_name,
            merged_range_lookup=merged_ranges_by_sheet,
            add_issue=add_issue,
        )

        if max_row > options.max_rows:
            add_issue(
                ReaderIssueCode.WORKBOOK_TOO_MANY_ROWS,
                IssueSeverity.CRITICAL,
                "Worksheet exceeds max_rows",
                sheet_name=sheet_name,
                details={"limit": str(options.max_rows)},
            )
        if max_column > options.max_columns:
            add_issue(
                ReaderIssueCode.WORKBOOK_TOO_MANY_COLUMNS,
                IssueSeverity.CRITICAL,
                "Worksheet exceeds max_columns",
                sheet_name=sheet_name,
                details={"limit": str(options.max_columns)},
            )

        if rows_to_read == 0:
            return _ParsedSheetResult(
                sheet=Sheet(
                    name=sheet_name,
                    index=sheet_index,
                    visibility=visibility,
                    max_row=max_row,
                    max_column=max_column,
                    merged_ranges=tuple(merged_ranges),
                    rows=(),
                ),
                rows_read=0,
                rows_skipped=max_row,
                cells_read=0,
                formula_cells=0,
                error_cells=0,
                stopped=rows_read >= options.max_rows,
            )

        formula_rows = sheet.iter_rows(
            min_row=1,
            max_row=rows_to_read,
            min_col=1,
            max_col=min(max_column, options.max_columns),
            values_only=False,
        )
        cached_rows = cached_sheet.iter_rows(
            min_row=1,
            max_row=rows_to_read,
            min_col=1,
            max_col=min(max_column, options.max_columns),
            values_only=False,
        )

        parsed_rows = 0
        local_cells = 0
        local_formula_cells = 0
        local_error_cells = 0
        rows: list[Row] = []
        skipped_rows = 0

        for formula_row, cached_row in zip_longest(formula_rows, cached_rows, fillvalue=()):
            if not formula_row:
                parsed_rows += 1
                continue

            if self._is_timeout(started, options):
                add_issue(
                    ReaderIssueCode.TIMEOUT_EXCEEDED,
                    IssueSeverity.CRITICAL,
                    "XLSX reader timeout exceeded",
                    retryable=True,
                )
                skipped_rows += rows_to_read - parsed_rows
                return _ParsedSheetResult(
                    sheet=Sheet(
                        name=sheet_name,
                        index=sheet_index,
                        visibility=visibility,
                        max_row=max_row,
                        max_column=max_column,
                        merged_ranges=tuple(merged_ranges),
                        rows=tuple(rows),
                    ),
                    rows_read=parsed_rows,
                    rows_skipped=skipped_rows,
                    cells_read=local_cells,
                    formula_cells=local_formula_cells,
                    error_cells=local_error_cells,
                    stopped=True,
                )

            row_result = self._parse_row(
                formula_row=tuple(formula_row),
                cached_row=tuple(cached_row) if cached_row else (),
                sheet_name=sheet_name,
                options=options,
                cells_read=cells_read + local_cells,
                add_issue=add_issue,
            )
            parsed_rows += 1

            if row_result.row is not None:
                rows.append(row_result.row)
                local_cells += row_result.parsed_cells
                local_formula_cells += row_result.formula_cells
                local_error_cells += row_result.error_cells

            if row_result.stopped:
                skipped_rows += rows_to_read - parsed_rows
                return _ParsedSheetResult(
                    sheet=Sheet(
                        name=sheet_name,
                        index=sheet_index,
                        visibility=visibility,
                        max_row=max_row,
                        max_column=max_column,
                        merged_ranges=tuple(merged_ranges),
                        rows=tuple(rows),
                    ),
                    rows_read=parsed_rows,
                    rows_skipped=skipped_rows,
                    cells_read=local_cells,
                    formula_cells=local_formula_cells,
                    error_cells=local_error_cells,
                    stopped=True,
                )

            if rows_read + parsed_rows >= options.max_rows and parsed_rows < rows_to_read:
                add_issue(
                    ReaderIssueCode.WORKBOOK_TOO_MANY_ROWS,
                    IssueSeverity.CRITICAL,
                    "Workbook exceeds max_rows",
                    sheet_name=sheet_name,
                    details={"limit": str(options.max_rows)},
                )
                skipped_rows += rows_to_read - parsed_rows
                return _ParsedSheetResult(
                    sheet=Sheet(
                        name=sheet_name,
                        index=sheet_index,
                        visibility=visibility,
                        max_row=max_row,
                        max_column=max_column,
                        merged_ranges=tuple(merged_ranges),
                        rows=tuple(rows),
                    ),
                    rows_read=parsed_rows,
                    rows_skipped=skipped_rows,
                    cells_read=local_cells,
                    formula_cells=local_formula_cells,
                    error_cells=local_error_cells,
                    stopped=True,
                )

            if cells_read + local_cells >= options.max_cells and parsed_rows < rows_to_read:
                add_issue(
                    ReaderIssueCode.CELL_LIMIT_EXCEEDED,
                    IssueSeverity.CRITICAL,
                    "Workbook exceeds max_cells",
                    sheet_name=sheet_name,
                    details={"limit": str(options.max_cells)},
                )
                skipped_rows += rows_to_read - parsed_rows
                return _ParsedSheetResult(
                    sheet=Sheet(
                        name=sheet_name,
                        index=sheet_index,
                        visibility=visibility,
                        max_row=max_row,
                        max_column=max_column,
                        merged_ranges=tuple(merged_ranges),
                        rows=tuple(rows),
                    ),
                    rows_read=parsed_rows,
                    rows_skipped=skipped_rows,
                    cells_read=local_cells,
                    formula_cells=local_formula_cells,
                    error_cells=local_error_cells,
                    stopped=True,
                )

        return _ParsedSheetResult(
            sheet=Sheet(
                name=sheet_name,
                index=sheet_index,
                visibility=visibility,
                max_row=max_row,
                max_column=max_column,
                merged_ranges=tuple(merged_ranges),
                rows=tuple(rows),
            ),
            rows_read=parsed_rows,
            rows_skipped=max_row - rows_to_read,
            cells_read=local_cells,
            formula_cells=local_formula_cells,
            error_cells=local_error_cells,
            stopped=False,
        )

    def _parse_row(
        self,
        *,
        formula_row: tuple[Any, ...],
        cached_row: tuple[Any, ...],
        sheet_name: str,
        options: ReaderOptions,
        cells_read: int,
        add_issue: Any,
    ) -> _ParsedRowResult:
        first_cell = next(
            (
                cell
                for cell in formula_row
                if cell is not None
                and getattr(cell, "value", None) is not None
                and getattr(cell, "column", None) is not None
                and getattr(cell, "row", None) is not None
            ),
            None,
        )
        if first_cell is None:
            return _ParsedRowResult(row=None, parsed_cells=0, formula_cells=0, error_cells=0)

        row_index = int(first_cell.row)
        cells: list[Cell] = []
        parsed_cells = 0
        formula_cells = 0
        error_cells = 0

        try:
            row_dimensions = first_cell.parent.row_dimensions[row_index]
            row_hidden = bool(getattr(row_dimensions, "hidden", False))
            raw_height = getattr(row_dimensions, "height", None)
            row_height = Decimal(str(raw_height)) if raw_height is not None else None
        except Exception:
            row_hidden = False
            row_height = None

        for formula_cell, cached_cell in zip_longest(formula_row, cached_row, fillvalue=None):
            if cells_read + parsed_cells >= options.max_cells:
                add_issue(
                    ReaderIssueCode.CELL_LIMIT_EXCEEDED,
                    IssueSeverity.CRITICAL,
                    "Workbook exceeds max_cells",
                    sheet_name=sheet_name,
                    row_index=row_index,
                    details={"limit": str(options.max_cells)},
                )
                return _ParsedRowResult(
                    row=Row(
                        index=row_index,
                        cells=tuple(cells),
                        hidden=row_hidden,
                        height=row_height,
                    )
                    if cells
                    else None,
                    parsed_cells=parsed_cells,
                    formula_cells=formula_cells,
                    error_cells=error_cells,
                    stopped=True,
                )

            if formula_cell is None or getattr(formula_cell, "value", None) is None:
                continue

            raw_column = getattr(formula_cell, "column", None)
            if isinstance(raw_column, str):
                try:
                    if raw_column.isalpha():
                        column_index = _decode_cell_reference(f"{raw_column}1")[0]
                    else:
                        column_index = int(raw_column)
                except ValueError:
                    continue
            elif isinstance(raw_column, int):
                column_index = raw_column
            else:
                continue

            if column_index <= 0:
                continue

            if column_index > options.max_columns:
                add_issue(
                    ReaderIssueCode.WORKBOOK_TOO_MANY_COLUMNS,
                    IssueSeverity.CRITICAL,
                    "Row exceeds max_columns",
                    sheet_name=sheet_name,
                    row_index=row_index,
                    cell_coordinate=CellCoordinate(row_index, column_index),
                    details={"limit": str(options.max_columns)},
                )
                return _ParsedRowResult(
                    row=Row(
                        index=row_index,
                        cells=tuple(cells),
                        hidden=row_hidden,
                        height=row_height,
                    )
                    if cells
                    else None,
                    parsed_cells=parsed_cells,
                    formula_cells=formula_cells,
                    error_cells=error_cells,
                    stopped=True,
                )

            try:
                parsed_cell = self._parse_cell(
                    formula_cell=formula_cell,
                    cached_cell=cached_cell,
                    column_index=column_index,
                    sheet_name=sheet_name,
                    row_index=row_index,
                    add_issue=add_issue,
                )
            except ValueError as exc:
                add_issue(
                    ReaderIssueCode.ROW_READ_FAILED,
                    IssueSeverity.ERROR,
                    "XLSX row has malformed cell",
                    sheet_name=sheet_name,
                    row_index=row_index,
                    details={"error": type(exc).__name__},
                )
                return _ParsedRowResult(
                    row=Row(
                        index=row_index,
                        cells=tuple(cells),
                        hidden=row_hidden,
                        height=row_height,
                    )
                    if cells
                    else None,
                    parsed_cells=parsed_cells,
                    formula_cells=formula_cells,
                    error_cells=error_cells,
                    stopped=False,
                )

            if parsed_cell is None:
                continue

            cells.append(parsed_cell)
            parsed_cells += 1
            formula_cells += parsed_cell.value_type is ValueType.FORMULA
            error_cells += parsed_cell.value_type is ValueType.ERROR

        if not cells:
            return _ParsedRowResult(row=None, parsed_cells=0, formula_cells=0, error_cells=0)
        return _ParsedRowResult(
            row=Row(index=row_index, cells=tuple(cells), hidden=row_hidden, height=row_height),
            parsed_cells=parsed_cells,
            formula_cells=formula_cells,
            error_cells=error_cells,
            stopped=False,
        )

    def _parse_cell(
        self,
        *,
        formula_cell: Any,
        cached_cell: Any,
        column_index: int,
        sheet_name: str,
        row_index: int,
        add_issue: Any,
    ) -> Cell | None:
        coordinate = CellCoordinate(formula_cell.row, column_index)
        data_type = str(getattr(formula_cell, "data_type", "n"))
        raw_value = getattr(formula_cell, "value", None)

        if data_type == "f":
            if not isinstance(raw_value, str):
                raise ValueError("Formula cells must store text formula")
            cached_value = _coerce_raw_value(getattr(cached_cell, "value", None))
            formula = Formula(formula_text=raw_value, cached_result=cached_value)
            add_issue(
                ReaderIssueCode.FORMULA_PRESENT,
                IssueSeverity.INFO,
                "Formula stored as raw text; cached value is not executed",
                sheet_name=sheet_name,
                row_index=row_index,
                cell_coordinate=coordinate,
            )
            error_code = _map_error_code(raw_value=cached_value)
            if error_code is not None:
                add_issue(
                    ReaderIssueCode.FORMULA_ERROR,
                    IssueSeverity.WARNING,
                    "Formula cached value is an error token",
                    sheet_name=sheet_name,
                    row_index=row_index,
                    cell_coordinate=coordinate,
                    details={"errorCode": error_code},
                )
            return Cell(
                coordinate=coordinate,
                row_index=coordinate.row,
                column_index=coordinate.column,
                value_type=ValueType.FORMULA,
                raw_value=raw_value,
                display_value=_display_value(cached_value),
                formula=formula,
                cached_value=cached_value,
                error_code=error_code,
            )

        if data_type == "e":
            if not isinstance(raw_value, str):
                raise ValueError("Malformed error token in error cell")
            error_code = _map_error_code(raw_value=raw_value) or "FORMULA_ERROR"
            add_issue(
                ReaderIssueCode.CELL_ERROR,
                IssueSeverity.ERROR,
                "Cell contains an error token",
                sheet_name=sheet_name,
                row_index=row_index,
                cell_coordinate=coordinate,
                details={"errorCode": error_code},
            )
            return Cell(
                coordinate=coordinate,
                row_index=coordinate.row,
                column_index=coordinate.column,
                value_type=ValueType.ERROR,
                raw_value=raw_value,
                display_value=raw_value,
                cached_value=None,
                error_code=error_code,
            )

        converted = _coerce_raw_value(raw_value)
        if converted is None:
            return None
        return Cell(
            coordinate=coordinate,
            row_index=coordinate.row,
            column_index=coordinate.column,
            value_type=_value_type_from_raw(converted),
            raw_value=converted,
            display_value=_display_value(converted),
        )

    def _parse_merged_ranges(
        self,
        *,
        sheet: Any,
        sheet_name: str,
        merged_range_lookup: dict[str, tuple[str, ...]] | None = None,
        add_issue: Any,
    ) -> list[MergedRange]:
        merged_ranges: list[MergedRange] = []
        worksheet_path = getattr(sheet, "_worksheet_path", None)
        merged_ranges_raw: tuple[Any, ...] | list[Any] = ()
        if worksheet_path is not None and merged_range_lookup is not None:
            merged_ranges_raw = list(merged_range_lookup.get(str(worksheet_path), ()))
        if hasattr(sheet, "merged_cells") and hasattr(sheet.merged_cells, "ranges"):
            merged_ranges_raw = sheet.merged_cells.ranges
        for merged in merged_ranges_raw:
            try:
                if isinstance(merged, str):
                    if ":" not in merged:
                        raise ValueError("Merged range should use A1:B2 format")
                    start, end = merged.split(":", 1)
                    min_col, min_row = _decode_cell_reference(start)
                    max_col, max_row = _decode_cell_reference(end)
                else:
                    min_row = int(merged.min_row)
                    min_col = int(merged.min_col)
                    max_row = int(merged.max_row)
                    max_col = int(merged.max_col)
            except Exception as exc:
                add_issue(
                    ReaderIssueCode.MERGED_RANGE_INVALID,
                    IssueSeverity.ERROR,
                    "Merged range is invalid",
                    sheet_name=sheet_name,
                    details={"error": type(exc).__name__},
                )
                continue

            try:
                merged_ranges.append(
                    MergedRange(
                        start_cell=CellCoordinate(min_row, min_col),
                        end_cell=CellCoordinate(max_row, max_col),
                    )
                )
            except ValueError as exc:
                add_issue(
                    ReaderIssueCode.MERGED_RANGE_INVALID,
                    IssueSeverity.ERROR,
                    "Merged range is invalid",
                    sheet_name=sheet_name,
                    details={"error": str(exc)},
                )
        return merged_ranges

    def _collect_merged_ranges(self, *, source: Any) -> dict[str, tuple[str, ...]]:
        merged_ranges_by_sheet: dict[str, tuple[str, ...]] = {}
        try:
            with ZipFile(source) as archive:
                for name in archive.namelist():
                    if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                        continue
                    content = archive.read(name)
                    try:
                        root = ET.fromstring(content)
                    except ET.ParseError:
                        continue
                    refs: list[str] = []
                    for node in root.iter():
                        tag = node.tag.rsplit("}", 1)[-1]
                        if tag == "mergeCell":
                            ref = node.attrib.get("ref")
                            if ref:
                                refs.append(ref)
                    if refs:
                        merged_ranges_by_sheet[name] = tuple(refs)
        except Exception:
            return {}
        return merged_ranges_by_sheet

    def _copy_source_to_temp(
        self,
        *,
        source: BinaryIO,
        destination: Any,
        options: ReaderOptions,
        started: float,
        add_issue: Any,
    ) -> int | None:
        bytes_read = 0
        while True:
            if self._is_timeout(started, options):
                add_issue(
                    ReaderIssueCode.TIMEOUT_EXCEEDED,
                    IssueSeverity.CRITICAL,
                    "XLSX reader timeout exceeded",
                    retryable=True,
                )
                return None

            chunk = source.read(_CHUNK_SIZE)
            if not chunk:
                break

            bytes_read += len(chunk)
            if bytes_read > options.max_file_size:
                add_issue(
                    ReaderIssueCode.FILE_TOO_LARGE,
                    IssueSeverity.CRITICAL,
                    "Workbook exceeds max_file_size",
                    details={"limitBytes": str(options.max_file_size)},
                )
                return None
            if bytes_read > options.memory_limit:
                add_issue(
                    ReaderIssueCode.MEMORY_LIMIT_EXCEEDED,
                    IssueSeverity.CRITICAL,
                    "Reader memory budget exceeded",
                    details={"limitBytes": str(options.memory_limit)},
                )
                return None
            destination.write(chunk)
        destination.flush()
        return bytes_read

    @staticmethod
    def _sheet_visibility(raw_state: str) -> SheetVisibility:
        if raw_state == "hidden":
            return SheetVisibility.HIDDEN
        if raw_state == "veryHidden":
            return SheetVisibility.VERY_HIDDEN
        return SheetVisibility.VISIBLE

    @staticmethod
    def _is_timeout(started: float, options: ReaderOptions) -> bool:
        return (time.monotonic() - started) > options.timeout_seconds

    @staticmethod
    def _elapsed(started: float) -> float:
        return max(0.0, time.monotonic() - started)

    @staticmethod
    def _statistics(*, bytes_read: int, started: float) -> ReaderStatistics:
        return ReaderStatistics(
            bytes_read=bytes_read,
            duration_seconds=max(0.0, time.monotonic() - started),
        )


def _coerce_raw_value(value: object) -> RawValue:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return value
    raise ValueError(f"Unsupported cell value type: {type(value)!r}")


def _value_type_from_raw(value: RawValue) -> ValueType:
    if value is None:
        return ValueType.EMPTY
    if isinstance(value, bool):
        return ValueType.BOOLEAN
    if isinstance(value, int):
        return ValueType.INTEGER
    if isinstance(value, Decimal):
        return ValueType.DECIMAL
    if isinstance(value, datetime):
        return ValueType.DATETIME
    if isinstance(value, date):
        return ValueType.DATE
    if isinstance(value, str):
        return ValueType.STRING
    raise ValueError(f"Unsupported raw value: {type(value)!r}")


def _decode_cell_reference(value: str) -> tuple[int, int]:
    match = _CELL_REFERENCE_PATTERN.fullmatch(value.upper())
    if match is None:
        raise ValueError(f"Неверный cell reference: {value!r}")
    column_name = match.group("column")
    row = int(match.group("row"))
    column = 0
    for char in column_name:
        column = column * 26 + (ord(char) - ord("A") + 1)
    return column, row


def _display_value(value: RawValue) -> str | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _map_error_code(*, raw_value: object) -> str | None:
    if not isinstance(raw_value, str) or not raw_value.startswith("#"):
        return None
    return _FORMULA_ERROR_CODE_BY_TOKEN.get(raw_value, "FORMULA_ERROR")
