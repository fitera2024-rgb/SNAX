from __future__ import annotations

from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

import snax_import.adapters.workbook.xlsx_reader as xlsx_reader_module
from snax_import.adapters.workbook import XlsxWorkbookReader
from snax_import.domain.raw_workbook import SheetVisibility, ValueType
from snax_import.ports.workbook_reader import ReaderIssueCode, ReaderOptions


def _set_formula_cache(raw: bytes, cell_ref: str, formula: str, cached: str) -> bytes:
    with ZipFile(BytesIO(raw), "r") as zin:
        sheet_xml = zin.read("xl/worksheets/sheet1.xml").decode("utf-8")
        old = f'<c r="{cell_ref}"><f>{formula}</f><v /></c>'
        new = f'<c r="{cell_ref}"><f>{formula}</f><v>{cached}</v></c>'
        if old not in sheet_xml:
            raise AssertionError(f"Unexpected formula cell layout for {cell_ref}")
        sheet_xml = sheet_xml.replace(old, new, 1)

        out = BytesIO()
        with ZipFile(out, "w") as zout:
            for name in zin.namelist():
                content = (
                    sheet_xml.encode("utf-8")
                    if name == "xl/worksheets/sheet1.xml"
                    else zin.read(name)
                )
                zout.writestr(name, content)
        return out.getvalue()


def _build_sample_workbook() -> bytes:
    workbook = Workbook()
    visible = workbook.active
    visible.title = "Visible"
    visible["A1"] = "001234"
    visible["B1"] = 10
    visible["B2"] = "=B1*2"
    visible.merge_cells("C1:E1")
    visible["C3"] = "#N/A"
    visible["C3"].data_type = "e"

    visible["D2"] = "#REF!"
    visible["D2"].data_type = "e"
    visible["E2"] = "#VALUE!"
    visible["E2"].data_type = "e"
    visible["F2"] = "#DIV/0!"
    visible["F2"].data_type = "e"

    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "Hidden value"

    buffer = BytesIO()
    workbook.save(buffer)
    return _set_formula_cache(buffer.getvalue(), cell_ref="B2", formula="B1*2", cached="20")


def _build_formula_error_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Errors"
    sheet["A1"] = 1
    sheet["A2"] = "#REF!"
    sheet["A2"].data_type = "e"
    sheet["B2"] = "#VALUE!"
    sheet["B2"].data_type = "e"
    sheet["C2"] = "#DIV/0!"
    sheet["C2"].data_type = "e"
    sheet["D1"] = "=A1+2"
    buffer = BytesIO()
    workbook.save(buffer)
    return _set_formula_cache(buffer.getvalue(), cell_ref="D1", formula="A1+2", cached="3")


def test_xlsx_reader_supports_xlsx_media_and_extensions_only() -> None:
    reader = XlsxWorkbookReader()

    assert reader.supports(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"
    )
    assert reader.supports(
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8"
    )
    assert reader.supports(extension=".XLSM")
    assert not reader.supports("application/pdf")
    assert not reader.supports("application/pdf", ".xlsx")
    assert not reader.supports()


def test_xlsx_reader_reads_hidden_sheet_formulas_error_cells_and_merged_ranges() -> None:
    reader = XlsxWorkbookReader()
    result = reader.read(BytesIO(_build_sample_workbook()), ReaderOptions())

    assert result.workbook is not None
    assert result.workbook.format.value == "XLSX"
    assert len(result.workbook.sheets) == 2

    visible = result.workbook.sheets[0]
    hidden = result.workbook.sheets[1]
    assert visible.visibility is SheetVisibility.VISIBLE
    assert hidden.visibility is SheetVisibility.HIDDEN
    assert visible.max_row >= 3
    assert len(visible.merged_ranges) == 1
    assert visible.rows[0].cells[0].raw_value == "001234"
    assert visible.rows[0].cells[0].value_type is ValueType.STRING

    formula_cell = next(
        cell for row in visible.rows for cell in row.cells if cell.value_type is ValueType.FORMULA
    )
    assert formula_cell.formula is not None
    assert formula_cell.formula.formula_text == "=B1*2"
    assert formula_cell.raw_value == "=B1*2"
    assert formula_cell.cached_value == 20
    assert formula_cell.display_value == "20"

    assert any(issue.code is ReaderIssueCode.FORMULA_PRESENT for issue in result.warnings)
    assert any(issue.code is ReaderIssueCode.CELL_ERROR for issue in result.errors)


def test_xlsx_reader_preserves_error_tokens_with_stable_codes() -> None:
    result = XlsxWorkbookReader().read(BytesIO(_build_formula_error_workbook()), ReaderOptions())

    assert result.workbook is not None
    sheet = result.workbook.sheets[0]
    cells_by_value = {
        cell.raw_value: cell.error_code
        for row in sheet.rows
        for cell in row.cells
        if cell.value_type is ValueType.ERROR
    }

    assert cells_by_value["#REF!"] == "FORMULA_ERROR_REF"
    assert cells_by_value["#VALUE!"] == "FORMULA_ERROR_VALUE"
    assert cells_by_value["#DIV/0!"] == "FORMULA_ERROR_DIV_ZERO"

    assert any(issue.code is ReaderIssueCode.CELL_ERROR for issue in result.errors)
    assert result.statistics.error_cells == 3


def test_xlsx_reader_stops_on_cells_limit_with_error_issue() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 1
    sheet["B1"] = 2
    sheet["C1"] = 3
    buffer = BytesIO()
    workbook.save(buffer)

    result = XlsxWorkbookReader().read(BytesIO(buffer.getvalue()), ReaderOptions(max_cells=2))

    assert not result.success
    assert any(issue.code is ReaderIssueCode.CELL_LIMIT_EXCEEDED for issue in result.errors)
    assert result.statistics.cells_read == 2


def test_xlsx_reader_preserves_blank_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "first"
    sheet["A3"] = "third"
    buffer = BytesIO()
    workbook.save(buffer)

    result = XlsxWorkbookReader().read(BytesIO(buffer.getvalue()), ReaderOptions())

    assert result.success
    assert result.workbook is not None
    rows = result.workbook.sheets[0].rows
    assert [row.index for row in rows] == [1, 2, 3]
    assert rows[1].cells == ()


def test_xlsx_reader_enforces_file_row_column_and_memory_limits() -> None:
    workbook = Workbook()
    first = workbook.active
    first["A1"] = "a"
    first["A2"] = "b"
    first["C2"] = "c"
    second = workbook.create_sheet("Second")
    second["A1"] = "d"
    second["A2"] = "e"
    buffer = BytesIO()
    workbook.save(buffer)
    source = buffer.getvalue()

    too_large = XlsxWorkbookReader().read(BytesIO(source), ReaderOptions(max_file_size=1))
    too_many_rows = XlsxWorkbookReader().read(BytesIO(source), ReaderOptions(max_rows=3))
    too_many_columns = XlsxWorkbookReader().read(BytesIO(source), ReaderOptions(max_columns=2))
    decompression_limited = XlsxWorkbookReader().read(
        BytesIO(source), ReaderOptions(memory_limit=len(source) + 100)
    )

    assert any(issue.code is ReaderIssueCode.FILE_TOO_LARGE for issue in too_large.errors)
    assert any(
        issue.code is ReaderIssueCode.WORKBOOK_TOO_MANY_ROWS for issue in too_many_rows.errors
    )
    assert too_many_rows.statistics.rows_read == 3
    assert any(
        issue.code is ReaderIssueCode.WORKBOOK_TOO_MANY_COLUMNS for issue in too_many_columns.errors
    )
    assert any(
        issue.code is ReaderIssueCode.MEMORY_LIMIT_EXCEEDED
        for issue in decompression_limited.errors
    )


def test_xlsx_reader_timeout_returns_issue_instead_of_traceback(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    ticks = iter((0.0, 2.0, 2.0))
    monkeypatch.setattr(
        xlsx_reader_module.time,
        "monotonic",
        lambda: next(ticks, 2.0),
    )

    result = XlsxWorkbookReader().read(
        BytesIO(_build_sample_workbook()),
        ReaderOptions(timeout_seconds=1.0),
    )

    assert result.workbook is None
    assert any(issue.code is ReaderIssueCode.TIMEOUT_EXCEEDED for issue in result.errors)


def test_xlsx_reader_reads_large_synthetic_workbook() -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Large")
    for row_index in range(500):
        sheet.append([f"{row_index:06d}", *range(1, 10)])
    buffer = BytesIO()
    workbook.save(buffer)

    result = XlsxWorkbookReader().read(BytesIO(buffer.getvalue()), ReaderOptions())

    assert result.success
    assert result.workbook is not None
    assert result.statistics.rows_read == 500
    assert result.statistics.cells_read == 5_000
    assert result.workbook.sheets[0].rows[0].cells[0].raw_value == "000000"


def test_xlsx_reader_returns_issue_for_invalid_corrupted_and_unsupported_zip() -> None:
    workbook = Workbook()
    workbook.active["A1"] = "value"
    valid = BytesIO()
    workbook.save(valid)

    unsupported = BytesIO()
    with ZipFile(unsupported, "w") as archive:
        archive.writestr("document.txt", "not a workbook")

    results = (
        XlsxWorkbookReader().read(BytesIO(b"not-an-xlsx"), ReaderOptions()),
        XlsxWorkbookReader().read(BytesIO(b"PK\x03\x04"), ReaderOptions()),
        XlsxWorkbookReader().read(
            BytesIO(valid.getvalue()[: len(valid.getvalue()) // 2]), ReaderOptions()
        ),
        XlsxWorkbookReader().read(BytesIO(unsupported.getvalue()), ReaderOptions()),
    )

    for result in results:
        assert not result.success
        assert result.workbook is None
        assert any(issue.code is ReaderIssueCode.UNSUPPORTED_FORMAT for issue in result.errors)
